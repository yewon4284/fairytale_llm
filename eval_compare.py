"""
eval_compare.py
SETA(우리 파이프라인) vs 다른 가이드라인/세이프가드 모델들의 동화 SAFE/UNSAFE 판정 성능을
사람 라벨(human_eval)을 정답으로 놓고 비교한다.

입력 파일 두 종류 (둘 다 선택적이지만, 표에 따라 필요한 것이 다름):

  1) --comparison (xlsx 또는 csv)
     모든 모델의 예측을 한 줄에 모아놓은 표. 최소 컬럼: filename, human_eval, 그리고
     모델별 예측 컬럼(들) (예: SETA, 라마가드4, 에스가드, ...).
     값은 대소문자 무관하게 "safe"/"unsafe"로 취급한다.

  2) --detail (jsonl)
     SETA(우리 시스템)의 상세 판정 근거. 한 줄에 하나씩:
     {"file", "title", "ground_truth", "predicted", "decision_source"(solar|kanana_force),
      "forced_categories", "average_score", "min_score", "flagged_categories", "fail_reasons", ...}
     FP/FN 원인 분해와 threshold 민감도 분석에 쓴다.

사용법:
    # 모델별 accuracy/precision/recall/F1 표
    python eval_compare.py --comparison "eval_comparison_final (1).xlsx" --summary

    # SETA vs 각 모델 McNemar exact 검정 (같은 421편에 대한 paired 비교)
    python eval_compare.py --comparison "eval_comparison_final (1).xlsx" --mcnemar

    # SETA가 틀린 이유 분해 (강제unsafe / 경계선 / 진짜낮은점수) + FN 목록
    python eval_compare.py --detail kanana_solar_eval.jsonl --breakdown

    # 합격선(threshold) 후보별로 FP가 몇 건 줄고 TP가 몇 건 희생되는지 시뮬레이션
    python eval_compare.py --detail kanana_solar_eval.jsonl --threshold-sweep

    # comparison의 human_eval과 detail의 ground_truth가 파일 단위로 일치하는지 확인
    python eval_compare.py --comparison "eval_comparison_final (1).xlsx" --detail kanana_solar_eval.jsonl --consistency

    # 위 다섯 개를 한 번에
    python eval_compare.py --comparison "eval_comparison_final (1).xlsx" --detail kanana_solar_eval.jsonl --all

GPU/모델 로딩 불필요 — 이미 나온 결과 파일들을 분석만 한다.
"""

import argparse
import csv
import json
import sys
from math import comb


# ── 파일 로더 ────────────────────────────────────────────────────────────────

def load_comparison_table(path):
    """xlsx 또는 csv를 읽어 (rows: List[dict], headers: List[str])로 반환.
    빈(None) 헤더 컬럼은 건너뛴다."""
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        raw_headers = [c.value for c in ws[1]]
        headers = [h for h in raw_headers if h is not None]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for h, v in zip(raw_headers, row):
                if h is not None:
                    d[h] = v
            rows.append(d)
        return rows, headers
    elif path.lower().endswith(".csv"):
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = [dict(r) for r in reader]
            headers = [h for h in (reader.fieldnames or []) if h is not None]
        return rows, headers
    else:
        raise ValueError(f"지원하지 않는 파일 형식(.xlsx 또는 .csv만 지원): {path}")


def load_detail_jsonl(path):
    recs = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                recs.append(json.loads(line))
    return recs


def norm_label(v):
    """safe/unsafe 판정 문자열을 대소문자 무관 대문자로 정규화. None/빈값은 그대로 반환."""
    if v is None:
        return None
    s = str(v).strip().upper()
    return s or None


# ── 통계 함수 ────────────────────────────────────────────────────────────────

def confusion_metrics(rows, pred_col, gt_col, positive="UNSAFE"):
    """rows(List[dict])에서 pred_col vs gt_col(둘 다 safe/unsafe류 라벨)로
    confusion matrix + accuracy/precision/recall/F1/specificity/FPR 계산.
    positive 라벨과 다른 값은 전부 '음성'으로 취급 (즉 라벨이 safe/unsafe 두 종류가
    아니어도 안전하게 동작 — 오탈자/결측은 음성 쪽으로 묶임)."""
    tp = fp = tn = fn = 0
    skipped = 0
    for r in rows:
        gt = norm_label(r.get(gt_col))
        pred = norm_label(r.get(pred_col))
        if gt is None or pred is None:
            skipped += 1
            continue
        gt_pos = (gt == positive)
        pred_pos = (pred == positive)
        if pred_pos and gt_pos:
            tp += 1
        elif pred_pos and not gt_pos:
            fp += 1
        elif not pred_pos and gt_pos:
            fn += 1
        else:
            tn += 1
    n = tp + fp + tn + fn
    acc = (tp + tn) / n if n else float("nan")
    prec = tp / (tp + fp) if (tp + fp) else float("nan")
    rec = tp / (tp + fn) if (tp + fn) else float("nan")
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) and prec == prec and rec == rec else float("nan")
    spec = tn / (tn + fp) if (tn + fp) else float("nan")
    fpr = fp / (fp + tn) if (fp + tn) else float("nan")
    return dict(n=n, skipped=skipped, tp=tp, fp=fp, tn=tn, fn=fn,
                acc=acc, prec=prec, rec=rec, f1=f1, spec=spec, fpr=fpr)


def mcnemar_exact(vec_a, vec_b):
    """두 분류기의 paired 정답여부(bool) 벡터를 받아 exact McNemar 검정.
    반환: (b, c, p) — b=A만 맞음 개수, c=B만 맞음 개수, p=양측 exact p값.
    scipy 없이 이항계수(math.comb)로 직접 계산 (표본 수 ~수백 수준이라 충분히 빠름)."""
    b = sum(1 for a, bb in zip(vec_a, vec_b) if a and not bb)
    c = sum(1 for a, bb in zip(vec_a, vec_b) if not a and bb)
    n = b + c
    if n == 0:
        return b, c, 1.0
    k = min(b, c)
    cum = sum(comb(n, i) for i in range(0, k + 1)) / (2 ** n)
    p = min(1.0, 2 * cum)
    return b, c, p


# ── 표 1: 모델별 accuracy/precision/recall/F1 ─────────────────────────────────

def print_summary_table(rows, headers, gt_col, filename_col, skip_cols):
    model_cols = [h for h in headers if h not in skip_cols and h != gt_col and h != filename_col]
    if not model_cols:
        print(f"\n비교할 모델 컬럼이 없습니다. headers={headers}")
        return

    print("\n" + "=" * 108)
    print(f"  모델별 성능 비교 — 정답: {gt_col} (양성 클래스 = UNSAFE)")
    print("=" * 108)
    print(f"  {'모델':50s} {'n':>4s} {'TP':>4s} {'FP':>4s} {'TN':>4s} {'FN':>4s} "
          f"{'정확도':>7s} {'정밀도':>7s} {'재현율':>7s} {'F1':>7s} {'특이도':>7s} {'FPR':>7s}")
    print(f"  {'-'*50} {'-'*4} {'-'*4} {'-'*4} {'-'*4} {'-'*4} "
          f"{'-'*7} {'-'*7} {'-'*7} {'-'*7} {'-'*7} {'-'*7}")
    for col in model_cols:
        d = confusion_metrics(rows, col, gt_col)
        skip_note = f" (라벨없음 {d['skipped']}건 제외)" if d["skipped"] else ""
        print(f"  {col:50s} {d['n']:4d} {d['tp']:4d} {d['fp']:4d} {d['tn']:4d} {d['fn']:4d} "
              f"{d['acc']*100:6.1f}% {d['prec']*100:6.1f}% {d['rec']*100:6.1f}% {d['f1']*100:6.1f}% "
              f"{d['spec']*100:6.1f}% {d['fpr']*100:6.1f}%{skip_note}")

    # 참고용: 그냥 항상 SAFE라고 찍었을 때 정확도 (trivial baseline)
    gt_vals = [norm_label(r.get(gt_col)) for r in rows]
    gt_vals = [v for v in gt_vals if v is not None]
    n_gt = len(gt_vals)
    n_safe = sum(1 for v in gt_vals if v != "UNSAFE")
    trivial_acc = n_safe / n_gt if n_gt else float("nan")
    print(f"\n  (참고) 무조건 SAFE라고 찍었을 때 정확도: {trivial_acc*100:.1f}%  "
          f"— 이보다 낮은 모델은 사실상 그냥 찍는 것보다 못한 판별력입니다.")
    print("=" * 108)


# ── 표 2: McNemar paired 검정 ──────────────────────────────────────────────

def print_mcnemar_table(rows, headers, gt_col, filename_col, skip_cols, our_col):
    model_cols = [h for h in headers if h not in skip_cols and h != gt_col and h != filename_col and h != our_col]
    if our_col not in headers:
        print(f"\n기준 모델 컬럼 '{our_col}'이 없습니다. headers={headers}")
        return
    if not model_cols:
        print("\n비교할 다른 모델이 없습니다.")
        return

    def correct_map(col):
        """filename -> 정답여부(bool) 딕셔너리. filename/gt/pred 중 하나라도 없으면 그 행은 제외.
        position(zip) 대신 filename으로 매칭해서, 모델별로 결측 위치가 달라도 항상 올바르게 짝지어지게 한다."""
        out = {}
        for r in rows:
            fn = r.get(filename_col)
            gt = norm_label(r.get(gt_col))
            pred = norm_label(r.get(col))
            if fn is None or gt is None or pred is None:
                continue
            out[fn] = (pred == gt)
        return out

    our_map = correct_map(our_col)

    print("\n" + "=" * 100)
    print(f"  McNemar exact 검정 — {our_col} vs 나머지 모델 (같은 데이터셋 paired 비교, filename으로 매칭)")
    print(f"  귀무가설: 두 모델의 정답률에 차이가 없다 (불일치 케이스 수만으로 판단, 표본크기 영향 안 받음)")
    print("=" * 100)
    print(f"  {'비교':45s} {our_col+'만 맞음':>12s} {'상대만 맞음':>12s} {'n(공통)':>8s} {'p(양측exact)':>14s}")
    for col in model_cols:
        col_map = correct_map(col)
        common_fn = sorted(set(our_map) & set(col_map))
        if len(common_fn) < len(our_map) or len(common_fn) < len(col_map):
            print(f"  (참고: {col}과 공통 파일 {len(common_fn)}건, {our_col}={len(our_map)}건/{col}={len(col_map)}건 중 일부만 매칭됨)")
        our_vec = [our_map[fn] for fn in common_fn]
        col_vec = [col_map[fn] for fn in common_fn]
        b, c, p = mcnemar_exact(our_vec, col_vec)
        sig = "*" if p < 0.05 else " "
        better = our_col if b > c else (col if c > b else "동률")
        print(f"  {our_col} vs {col:40s} {b:12d} {c:12d} {len(common_fn):8d} {p:13.4g}{sig}  -> {better}가 더 우세")
    print("\n  * = p<0.05 (통계적으로 유의한 차이)")
    print("=" * 100)


# ── 표 3: FP/FN 원인 분해 (detail jsonl 필요) ─────────────────────────────────

def print_breakdown(recs, pass_min=4.0):
    total = len(recs)
    fps = [r for r in recs if r.get("predicted") == "unsafe" and r.get("ground_truth") == "safe"]
    fns = [r for r in recs if r.get("predicted") == "safe" and r.get("ground_truth") == "unsafe"]

    print("\n" + "=" * 100)
    print(f"  FP/FN 원인 분해 (전체 {total}건, FP {len(fps)}건 / FN {len(fns)}건)")
    print("=" * 100)

    fp_forced = [r for r in fps if r.get("decision_source") == "kanana_force"]
    fp_solar = [r for r in fps if r.get("decision_source") == "solar"]
    fp_solar_border = [r for r in fp_solar if (r.get("average_score") or 0) >= pass_min]
    fp_solar_low = [r for r in fp_solar if (r.get("average_score") or 0) < pass_min]

    print(f"\n[FP {len(fps)}건 = 사람은 safe, 우리는 unsafe]")
    print(f"  ① 카나나 강제 unsafe(kanana_force): {len(fp_forced)}건 ({len(fp_forced)/len(fps)*100:.1f}%)" if fps else "  (FP 없음)")
    if fp_forced:
        from collections import Counter
        cat_counter = Counter(tuple(sorted(r.get("forced_categories") or [])) for r in fp_forced)
        for cats, cnt in cat_counter.most_common():
            print(f"      {cats}: {cnt}건")
    if fps:
        print(f"  ② Solar 판정, 평균 {pass_min}점 이상(경계선): {len(fp_solar_border)}건 ({len(fp_solar_border)/len(fps)*100:.1f}%)")
        print(f"  ③ Solar 판정, 평균 {pass_min}점 미만(진짜 낮은 점수): {len(fp_solar_low)}건 ({len(fp_solar_low)/len(fps)*100:.1f}%)")

    parse_issue = [r for r in recs if any("파싱" in x for x in (r.get("fail_reasons") or []))]
    fp_parse = [r for r in parse_issue if r in fps]
    if parse_issue:
        print(f"\n[참고] fail_reasons에 'JSON 파싱' 문제 언급된 레코드: {len(parse_issue)}건 "
              f"(그중 FP {len(fp_parse)}건) — 내용과 무관한 엔지니어링 이슈로 unsafe 처리됐을 가능성")

    print(f"\n[FN {len(fns)}건 = 사람은 unsafe, 우리는 safe로 통과 — 안전 관점에서는 이쪽이 더 위험]")
    for r in sorted(fns, key=lambda x: -(x.get("average_score") or 0)):
        flagged = r.get("flagged_categories") or []
        print(f"  {r.get('file',''):14s} {r.get('title',''):28s} avg={r.get('average_score')}  "
              f"flagged={flagged}  src={r.get('decision_source')}")
    print("=" * 100)


# ── 표 4: threshold 민감도 시뮬레이션 (detail jsonl 필요) ─────────────────────

def print_threshold_sweep(recs, thresholds=(4.5, 4.4, 4.3, 4.2, 4.1, 4.0), pass_min=4.0):
    """decision_source=solar 이면서 현재 predicted=unsafe인 것들(=avg 기준 미달로 떨어진 것들) 대상으로,
    average_score 합격선을 낮췄을 때 몇 건이 safe로 뒤집히는지, 그중 실제 safe(FP 해소, 좋음)와
    실제 unsafe(TP 손실, 나쁨)가 각각 몇 건인지 계산. min_score >= pass_min 조건은 그대로 유지한다고 가정."""
    solar_recs = [r for r in recs if r.get("decision_source") == "solar" and r.get("average_score") is not None]
    currently_unsafe = [r for r in solar_recs if r.get("predicted") == "unsafe"]
    fp_now = sum(1 for r in currently_unsafe if r.get("ground_truth") == "safe")
    tp_now = sum(1 for r in currently_unsafe if r.get("ground_truth") == "unsafe")

    print("\n" + "=" * 100)
    print(f"  합격선(average_score) 민감도 시뮬레이션 — 현재 4.5 기준, decision_source=solar 대상")
    print(f"  (카나나 강제 unsafe 44건은 이 threshold와 무관하므로 대상에서 제외)")
    print("=" * 100)
    print(f"  현재(4.5) 기준 solar-unsafe {len(currently_unsafe)}건 중 실제 safe(FP)={fp_now}건, 실제 unsafe(TP)={tp_now}건")
    print(f"\n  {'후보 threshold':>14s} {'FP 해소(safe로 correct 전환)':>28s} {'TP 손실(unsafe를 safe로 오판)':>28s} {'net(+가 이득)':>14s}")
    for thr in thresholds:
        flips = [r for r in currently_unsafe if r["average_score"] >= thr and (r.get("min_score") or 0) >= pass_min]
        fp_fixed = sum(1 for r in flips if r.get("ground_truth") == "safe")
        tp_lost = sum(1 for r in flips if r.get("ground_truth") == "unsafe")
        print(f"  {thr:14.1f} {fp_fixed:28d} {tp_lost:28d} {fp_fixed - tp_lost:14d}")
    print("\n  주의: min_score/신체안전체크 조건은 고정한 채 average_score 기준만 바꾼 시뮬레이션입니다.")
    print("  실제로 임계값을 낮추면 Solar가 매기는 점수 분포 자체가 달라질 수 있어(재채점 아님, 사후 시뮬레이션),")
    print("  참고용 상한선으로만 쓰세요 — 실제 재실행 결과와 비교해 검증할 것.")
    print("=" * 100)


# ── 표 5: 두 파일 간 라벨 일관성 체크 ────────────────────────────────────────

def print_consistency_check(comp_rows, filename_col, gt_col, detail_recs, detail_file_key="file"):
    comp_by_fn = {r.get(filename_col): r for r in comp_rows if r.get(filename_col)}
    detail_by_fn = {r.get(detail_file_key): r for r in detail_recs if r.get(detail_file_key)}

    only_comp = sorted(set(comp_by_fn) - set(detail_by_fn))
    only_detail = sorted(set(detail_by_fn) - set(comp_by_fn))
    common = sorted(set(comp_by_fn) & set(detail_by_fn))

    mismatches = []
    for fn in common:
        gt_c = norm_label(comp_by_fn[fn].get(gt_col))
        gt_d = norm_label(detail_by_fn[fn].get("ground_truth"))
        if gt_c != gt_d:
            mismatches.append((fn, comp_by_fn[fn].get("title", detail_by_fn[fn].get("title", "")), gt_c, gt_d))

    print("\n" + "=" * 100)
    print("  라벨 일관성 체크 — comparison 파일의 정답과 detail 파일의 ground_truth가 같은지")
    print("=" * 100)
    print(f"  comparison {len(comp_by_fn)}건 / detail {len(detail_by_fn)}건 / 공통 {len(common)}건")
    if only_comp:
        print(f"  comparison에만 있음: {len(only_comp)}건 {only_comp[:10]}{' ...' if len(only_comp) > 10 else ''}")
    if only_detail:
        print(f"  detail에만 있음: {len(only_detail)}건 {only_detail[:10]}{' ...' if len(only_detail) > 10 else ''}")
    print(f"\n  ground_truth 불일치: {len(mismatches)}건")
    for fn, title, gt_c, gt_d in mismatches:
        print(f"    {fn:14s} {title:28s} comparison={gt_c}  detail={gt_d}")
    if mismatches:
        print("\n  주의: 위 파일들은 두 파일 중 어느 라벨이 최신/정정본인지 확인 후 하나로 통일해서 써야")
        print("  성능 수치가 재현 가능하고 팀원 간 비교가 일관됩니다.")
    print("=" * 100)


# ── main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--comparison", default=None, help="filename/human_eval/모델들 컬럼이 있는 xlsx 또는 csv")
    parser.add_argument("--detail", default=None, help="SETA 상세 판정 근거 jsonl (kanana_solar_eval.jsonl 형식)")
    parser.add_argument("--gt-col", default="human_eval", help="comparison 파일에서 정답으로 쓸 컬럼명")
    parser.add_argument("--filename-col", default="filename", help="comparison 파일의 파일명 컬럼")
    parser.add_argument("--detail-file-key", default="file", help="detail jsonl의 파일명 키")
    parser.add_argument("--our-col", default="SETA", help="McNemar 비교 기준이 될 우리 시스템 컬럼명")
    parser.add_argument("--skip-cols", action="append", default=["filename", "isbn", "title", "agree"],
                         help="comparison 표에서 모델 컬럼으로 취급하지 않을 컬럼 (여러 번 지정 가능)")
    parser.add_argument("--pass-min", type=float, default=4.0, help="항목별 최저 합격선 (min_score 기준)")
    parser.add_argument("--thresholds", type=float, nargs="+", default=[4.5, 4.4, 4.3, 4.2, 4.1, 4.0],
                         help="--threshold-sweep에서 시험해볼 average_score 후보들")
    parser.add_argument("--summary", action="store_true", help="모델별 accuracy/precision/recall/F1 표")
    parser.add_argument("--mcnemar", action="store_true", help="our-col vs 나머지 모델 McNemar exact 검정")
    parser.add_argument("--breakdown", action="store_true", help="FP/FN 원인 분해 (--detail 필요)")
    parser.add_argument("--threshold-sweep", action="store_true", help="합격선 후보별 FP/TP 시뮬레이션 (--detail 필요)")
    parser.add_argument("--consistency", action="store_true", help="두 파일 간 라벨 일관성 체크 (둘 다 필요)")
    parser.add_argument("--all", action="store_true", help="가능한 표를 전부 출력")
    args = parser.parse_args()

    if not any([args.summary, args.mcnemar, args.breakdown, args.threshold_sweep, args.consistency, args.all]):
        print("표시할 표를 하나 이상 지정하세요 (--summary/--mcnemar/--breakdown/--threshold-sweep/--consistency/--all).")
        sys.exit(1)

    comp_rows, comp_headers = (None, None)
    if args.comparison:
        comp_rows, comp_headers = load_comparison_table(args.comparison)

    detail_recs = None
    if args.detail:
        detail_recs = load_detail_jsonl(args.detail)

    if args.summary or args.all:
        if comp_rows is None:
            print("\n--summary는 --comparison 파일이 필요합니다.")
        else:
            print_summary_table(comp_rows, comp_headers, args.gt_col, args.filename_col, set(args.skip_cols))

    if args.mcnemar or args.all:
        if comp_rows is None:
            print("\n--mcnemar는 --comparison 파일이 필요합니다.")
        else:
            print_mcnemar_table(comp_rows, comp_headers, args.gt_col, args.filename_col, set(args.skip_cols), args.our_col)

    if args.breakdown or args.all:
        if detail_recs is None:
            print("\n--breakdown은 --detail 파일이 필요합니다.")
        else:
            print_breakdown(detail_recs, pass_min=args.pass_min)

    if args.threshold_sweep or args.all:
        if detail_recs is None:
            print("\n--threshold-sweep은 --detail 파일이 필요합니다.")
        else:
            print_threshold_sweep(detail_recs, thresholds=tuple(args.thresholds), pass_min=args.pass_min)

    if args.consistency or args.all:
        if comp_rows is None or detail_recs is None:
            print("\n--consistency는 --comparison과 --detail 파일이 모두 필요합니다.")
        else:
            print_consistency_check(comp_rows, args.filename_col, args.gt_col, detail_recs, args.detail_file_key)


if __name__ == "__main__":
    main()
